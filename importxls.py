import openpyxl
import sqlite3

path = "slova.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

row = sheet_obj.max_row

con = sqlite3.connect("bot.db")
cur = con.cursor()
con.executescript("""
        BEGIN TRANSACTION;
        CREATE TABLE IF NOT EXISTS 'dictionary' (
            'id'	INTEGER NOT NULL UNIQUE,
            'en'	TEXT NOT NULL UNIQUE,
            'ru'	TEXT,
            'transcription'	TEXT,
            'voice'	TEXT,
            PRIMARY KEY('id' AUTOINCREMENT)
        );
        COMMIT;
        """)
con.commit

for i in range(1, row + 1): 
    cell_obj = sheet_obj.cell(row = i, column = 1) 
    print(cell_obj.value)
    res = con.execute('INSERT INTO dictionary (en) VALUES ("%s");'% cell_obj.value)
    con.commit()
      
con.close()